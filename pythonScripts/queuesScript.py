import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import utils
import logging

def process_queues():
    metadata_folder = 'queues'
    worksheet_name = metadata_folder
    processing_path = f"{utils.metadata_path}/{metadata_folder}"
    print(f"Processing directory_path {processing_path}")

    if not os.path.exists(processing_path):
        message = f"The directory {processing_path} does not exist. Skipping processing for {worksheet_name}.\n"
        logging.info(message)
        return

    wb = utils.open_workbook(utils.config_file_path)

    if worksheet_name in wb.sheetnames:
        wb.remove(wb[worksheet_name])

    ws = wb.create_sheet(title=worksheet_name)
    header = ['Name', 'DoesSendEmailToMembers', 'Email', 'QueueRoutingConfig', 'QueueSobject', 'PublicGroups', 'RoleAndSubordinates', 'RoleAndSubordinatesInternal', 'Roles', 'users']
    ws.append(header)
    wb.save(utils.config_file_path)

    for root_dir, dirs, files in os.walk(processing_path):
        for file in files:
            if file.endswith('queue-meta.xml'):
                tree = ET.parse(os.path.join(root_dir, file))
                root = tree.getroot()

                nameVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}name').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}name') is not None else ""
                doesSendEmaiToMembersVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}doesSendEmailToMembers').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}doesSendEmailToMembers') is not None else ""
                emailVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}email').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}email') is not None else ""
                queueRoutingConfigVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}queueRoutingConfig').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}queueRoutingConfig') is not None else ""
                sobjectVar = ", ".join([sobject.text for sobject in root.findall('.//{http://soap.sforce.com/2006/04/metadata}sobjectType')])
                publicGroupVar = ", ".join([group.text for group in root.findall('.//{http://soap.sforce.com/2006/04/metadata}publicGroup')])
                roleAndSubordinatesVar = ", ".join([role.text for role in root.findall('.//{http://soap.sforce.com/2006/04/metadata}queueRoleAndSubordinate')])
                roleAndSubordinatesInternalVar = ", ".join([role.text for role in root.findall('.//{http://soap.sforce.com/2006/04/metadata}queueRoleAndSubordinateInternal')])
                rolesVar = ", ".join([role.text for role in root.findall('.//{http://soap.sforce.com/2006/04/metadata}role')])
                usersVar = ", ".join([user.text for user in root.findall('.//{http://soap.sforce.com/2006/04/metadata}user')])

                ws.append([nameVar, doesSendEmaiToMembersVar, emailVar, queueRoutingConfigVar, sobjectVar, publicGroupVar, roleAndSubordinatesVar, roleAndSubordinatesInternalVar, rolesVar, usersVar])

    if ws.max_row > 1:
        tab = Table(displayName=worksheet_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for row in ws.iter_rows():
            for cell in row:
                ws[str(cell.coordinate)].font = Font(size=9)
            for cell in ws["1:1"]:
                cell.font = Font(color='FFFFFF', size=9)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 0.98

        wb.save(utils.config_file_path)
        print(f"Workbook Saved: {utils.config_file_path}")
    else:
        print(f"No current data for this metadata type in {utils.project_path}")
    print(f"Processing complete: {worksheet_name}\n")

process_queues()
