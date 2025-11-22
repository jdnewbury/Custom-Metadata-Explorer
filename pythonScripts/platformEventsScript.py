import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import utils
import logging

def process_platform_events():
    metadata_folder = 'objects'
    worksheet_name = 'platformEvents'
    processing_path = f"{utils.metadata_path}/{metadata_folder}"
    print(f"Processing directory_path {processing_path}")

    if not os.path.exists(processing_path):
        message = f"The directory {processing_path} does not exist. Skipping processing for {worksheet_name}.\n"
        logging.info(message)
        return

    wb = utils.open_workbook(utils.config_file_path)

    if worksheet_name in wb.sheetnames:
        wb.remove(wb[worksheet_name])

    ws = wb.create_sheet(title=worksheet_name)
    header = ['Object', 'APIName', 'UniqueKey', 'Label', 'FieldType', 'Length', 'ExternalId', 'Required', 'FeedHistoryTracking', 'HistoryTracking', 'ReferenceTo', 'RelationshipName']
    ws.append(header)
    wb.save(utils.config_file_path)

    for root_dir, dirs, files in os.walk(processing_path):
        for dir_name in dirs:
            if dir_name.endswith('__e'):
                fields_dir_path = os.path.join(root_dir, dir_name, 'fields')
                if os.path.exists(fields_dir_path):

                    for file in os.listdir(fields_dir_path):
                        #print("file:", file)
                        file_path = os.path.join(fields_dir_path, file)

                        api_object_name = os.path.basename(os.path.dirname(os.path.dirname(file_path)))

                        tree = ET.parse(file_path)
                        root = tree.getroot()

                        api_field_name = root.find(".//{http://soap.sforce.com/2006/04/metadata}fullName").text if root.find(".//{http://soap.sforce.com/2006/04/metadata}fullName") is not None else ''
                        row_data = [
                            api_object_name,
                            api_field_name,
                            f"{api_object_name}.{api_field_name}",
                            root.find(".//{http://soap.sforce.com/2006/04/metadata}label").text if root.find(".//{http://soap.sforce.com/2006/04/metadata}label") is not None else '',
                            root.find(".//{http://soap.sforce.com/2006/04/metadata}type").text if root.find(".//{http://soap.sforce.com/2006/04/metadata}type") is not None else '',
                            root.find(".//{http://soap.sforce.com/2006/04/metadata}length").text if root.find(".//{http://soap.sforce.com/2006/04/metadata}length") is not None else '',
                            root.find(".//{http://soap.sforce.com/2006/04/metadata}externalId").text if root.find(".//{http://soap.sforce.com/2006/04/metadata}externalId") is not None else '',
                            root.find(".//{http://soap.sforce.com/2006/04/metadata}required").text if root.find(".//{http://soap.sforce.com/2006/04/metadata}required") is not None else '',
                            root.find(".//{http://soap.sforce.com/2006/04/metadata}trackFeedHistory").text if root.find(".//{http://soap.sforce.com/2006/04/metadata}trackFeedHistory") is not None else '',
                            root.find(".//{http://soap.sforce.com/2006/04/metadata}trackHistory").text if root.find(".//{http://soap.sforce.com/2006/04/metadata}trackHistory") is not None else '',
                            root.find(".//{http://soap.sforce.com/2006/04/metadata}referenceTo").text if root.find(".//{http://soap.sforce.com/2006/04/metadata}referenceTo") is not None else '',
                            root.find(".//{http://soap.sforce.com/2006/04/metadata}relationshipName").text if root.find(".//{http://soap.sforce.com/2006/04/metadata}relationshipName") is not None else ''
                        ]

                        ws.append(row_data)

    if ws.max_row > 1:
        tab = Table(displayName=worksheet_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for row in ws.iter_rows():
            for cell in row:
                ws[str(cell.coordinate)].font = Font(size=9)
            for cell in ws["1:1"]:
                cell.font = Font(color='FFFFFF', size=9)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 0.98

        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15

        wb.save(utils.config_file_path)
        print(f"Workbook Saved: {utils.config_file_path}")
    else:
        print(f"No current data for this metadata type in {utils.project_path}")
    print(f"Processing complete: {worksheet_name}\n")

process_platform_events()
