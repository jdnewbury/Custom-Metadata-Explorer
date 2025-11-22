import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import utils
import logging

def process_reports():
    metadata_folder = 'reports'
    worksheet_name = metadata_folder
    processing_path = f"{utils.metadata_path}/{metadata_folder}"
    print(f"Processing directory_path {processing_path}")

    if not os.path.exists(processing_path):
        message = f"The directory {processing_path} does not exist. Skipping processing for {worksheet_name}.\n"
        logging.info(message)
        return

    wb = utils.open_workbook(utils.config_file_path)

    if worksheet_name in wb.sheetnames:
        wb.remove(wb[worksheet_name])

    ws = wb.create_sheet(title=worksheet_name)
    header = ['folder','name', 'reportType', 'format', 'scope', 'showDetails', 'showGrandTotal', 'showSubTotals', 'sortColumn', 'sortOrder', 'columns', 'timeFrameFilter','groupingsDown', 'params']
    ws.append(header)
    wb.save(utils.config_file_path)

    for root_dir, dirs, files in os.walk(processing_path):
        for file in files:
            file_path = os.path.join(root_dir, file)
            #print(f"Processing file: {file}")
            tree = ET.parse(file_path)
            root = tree.getroot()

            folderVar = metadata_folder
            nameVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}name').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}name') is not None else ""
            reportTypeVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}reportType').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}reportType') is not None else ""
            formatVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}format').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}format') is not None else ""
            scopeVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}scope').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}scope') is not None else ""
            showDetailsVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}showDetails').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}showDetails') is not None else ""
            showGrandTotalVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}showGrandTotal').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}showGrandTotal') is not None else ""
            showSubTotalsVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}showSubTotals').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}showSubTotals') is not None else ""
            sortColumnVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}sortColumn').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}sortColumn') is not None else ""
            sortOrderVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}sortOrder').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}sortOrder') is not None else ""

            columns = root.findall('.//{http://soap.sforce.com/2006/04/metadata}columns')
            i = 0
            columns_text = ""
            for column in columns:
                i += 1
                if i > 1:
                    columns_text += ', '
                columns_text += column.find('.//{http://soap.sforce.com/2006/04/metadata}field').text

            timeFrameFilter_text = ""
            if (root.find('.//{http://soap.sforce.com/2006/04/metadata}timeFrameFilter')) is not None:
                timeFrameFilter_text = f"dateColumn: {root.find('.//{http://soap.sforce.com/2006/04/metadata}dateColumn').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}dateColumn') is not None else ""}"
                timeFrameFilter_text += f", interval: {root.find('.//{http://soap.sforce.com/2006/04/metadata}interval').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}interval') is not None else ""}"
                if (startDate := root.find('.//{http://soap.sforce.com/2006/04/metadata}startDate')) is not None:
                    timeFrameFilter_text += f', startDate: {startDate.text}'

            groupingsDowns = root.findall('.//{http://soap.sforce.com/2006/04/metadata}groupingsDown')

            groupingsDown_text = ""
            j = 0
            for groupingsDown in groupingsDowns:
                j += 1
                if j > 1:
                    groupingsDown_text += '\n'
                if (dateGranularity := groupingsDown.find('.//{http://soap.sforce.com/2006/04/metadata}dateGranularity')) is not None:
                    groupingsDown_text += f'dateGranularity: {dateGranularity.text}'
                if (field := groupingsDown.find('.//{http://soap.sforce.com/2006/04/metadata}field')) is not None:
                    groupingsDown_text += f', field: {field.text}'
                if (sortOrder := groupingsDown.find('.//{http://soap.sforce.com/2006/04/metadata}sortOrder')) is not None:
                    groupingsDown_text += f', sortOrder: {sortOrder.text}'

            params = root.findall('.//{http://soap.sforce.com/2006/04/metadata}params')

            params_text = ""
            j = 0
            for param in params:
                j += 1
                if j > 1:
                    params_text += '\n'
                if (name := param.find('.//{http://soap.sforce.com/2006/04/metadata}name')) is not None:
                    params_text += f'name: {name.text}'
                if (value := param.find('.//{http://soap.sforce.com/2006/04/metadata}value')) is not None:
                    params_text += f', value: {value.text}'

            # Step 6
            ws.append([folderVar, nameVar, reportTypeVar, formatVar, scopeVar, showDetailsVar, showGrandTotalVar, showSubTotalsVar, sortColumnVar, sortOrderVar, columns_text, timeFrameFilter_text, groupingsDown_text, params_text])

        for cell in ws['M']:
            cell.alignment = Alignment(wrap_text=True)
        for cell in ws['N']:
            cell.alignment = Alignment(wrap_text=True)

    if ws.max_row > 1:
        tab = Table(displayName=worksheet_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for row in ws.iter_rows():
            for cell in row:
                ws[str(cell.coordinate)].font = Font(size=9)
            for cell in ws["1:1"]:
                cell.font = Font(color='FFFFFF', size=9)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 0.98

        ws.column_dimensions['K'].width = 120
        for cell in ws['K']:
            cell.alignment = Alignment(wrap_text=True)

        ws.column_dimensions['L'].width = 80
        for cell in ws['L']:
            cell.alignment = Alignment(wrap_text=True)
        ws.column_dimensions['M'].width = 60

        wb.save(utils.config_file_path)
        print(f"Workbook Saved: {utils.config_file_path}")
    else:
        print(f"No current data for this metadata type in {utils.project_path}")
    print(f"Processing complete: {worksheet_name}\n")

process_reports()
