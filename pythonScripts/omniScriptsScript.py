import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import utils
import logging

def process_omni_scripts():
    metadata_folder = 'omniScripts'
    worksheet_name = metadata_folder
    processing_path = f"{utils.metadata_path}/{metadata_folder}"
    print(f"Processing directory_path {processing_path}")

    if not os.path.exists(processing_path):
        message = f"The directory {processing_path} does not exist. Skipping processing for {worksheet_name}.\n"
        logging.info(message)
        return

    wb = utils.open_workbook(utils.config_file_path)

    if worksheet_name in wb.sheetnames:
        wb.remove(wb[worksheet_name])

    ws = wb.create_sheet(title=worksheet_name)
    header = [
        'name',
        'active',
        'description',
        'uniqueName',
        'omniProcessType',
        'type',
        'subType',
        'versionNumber',
        'isIntegrationProcedure',
        'isMetadataCacheDisabled',
        'isOmniScriptEmbeddable',
        'isTestProcedure',
        'isWebCompEnabled',
        'webComponentKey'
    ]
    ws.append(header)
    wb.save(utils.config_file_path)

    for root_dir, dirs, files in os.walk(processing_path):
        for file in files:
            file_path = os.path.join(root_dir, file)
            #print(f"Processing file: {file}")
            tree = ET.parse(file_path)
            root = tree.getroot()

            nameVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}name').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}name') is not None else ""
            activeVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}active').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}active') is not None else ""
            descriptionVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}description').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}description') is not None else ""
            uniqueNameVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}uniqueName').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}uniqueName') is not None else ""
            omniProcessTypeVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}omniProcessType').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}omniProcessType') is not None else ""
            typeVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}type').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}type') is not None else ""
            subTypeVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}subType').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}subType') is not None else ""
            versionNumberVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}versionNumber').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}versionNumber') is not None else ""
            isIntegrationProcedureVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}isIntegrationProcedure').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}isIntegrationProcedure') is not None else ""
            isMetadataCacheDisabledVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}isMetadataCacheDisabled').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}isMetadataCacheDisabled') is not None else ""
            isOmniScriptEmbeddableVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}isOmniScriptEmbeddable').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}isOmniScriptEmbeddable') is not None else ""
            isTestProcedureVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}isTestProcedure').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}isTestProcedure') is not None else ""
            isWebCompEnabledVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}isWebCompEnabled').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}isWebCompEnabled') is not None else ""
            webComponentKeyVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}webComponentKey').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}webComponentKey') is not None else ""

            ws.append([nameVar, activeVar, descriptionVar, uniqueNameVar, omniProcessTypeVar, typeVar, subTypeVar, versionNumberVar, isIntegrationProcedureVar, isMetadataCacheDisabledVar, isOmniScriptEmbeddableVar, isTestProcedureVar, isWebCompEnabledVar, webComponentKeyVar])

    if ws.max_row > 1:
        tab = Table(displayName=worksheet_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for row in ws.iter_rows():
            for cell in row:
                ws[str(cell.coordinate)].font = Font(size=9)
            for cell in ws["1:1"]:
                cell.font = Font(color='FFFFFF', size=9)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 0.98

        wb.save(utils.config_file_path)
        print(f"Workbook Saved: {utils.config_file_path}")
    else:
        print(f"No current data for this metadata type in {utils.project_path}")
    print(f"Processing complete: {worksheet_name}\n")

process_omni_scripts()
