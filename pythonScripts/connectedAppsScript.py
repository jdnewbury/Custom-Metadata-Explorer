import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import utils
import logging

def process_connected_apps():
    # Metadata folder to process
    metadata_folder = 'connectedApps'
    worksheet_name = metadata_folder

    # Construct the full path to the directory
    processing_path = f"{utils.metadata_path}/{metadata_folder}"
    print(f"Processing directory_path {processing_path}")

    # Check if the directory exists
    if not os.path.exists(processing_path):
        message = f"The directory {processing_path} does not exist. Skipping processing for {worksheet_name}.\n"
        logging.info(message)
        return

    # Get the config file
    wb = utils.open_workbook(utils.config_file_path)

    # Remove sheet if exists
    if worksheet_name in wb.sheetnames:
        wb.remove(wb[worksheet_name])

    # Create new sheet
    ws = wb.create_sheet(title=worksheet_name)

    # Create Header
    header = [
        'label',
        'contactEmail',
        'callbackUrl',
        'consumerKey',
        'scopes',
        'ipRelaxation',
        'refreshTokenPolicy',
        'permissionSetName'
    ]
    ws.append(header)
    wb.save(utils.config_file_path)

    # Loop through each file in the directory
    for root_dir, dirs, files in os.walk(processing_path):
        for file in files:
            if file.endswith('.xml'):
                file_path = os.path.join(root_dir, file)
                #print(f"Processing file: {file}")
                tree = ET.parse(file_path)
                root = tree.getroot()

                labelVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}label').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}label') is not None else ""
                contactEmailVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}contactEmail').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}contactEmail') is not None else ""
                callbackUrlVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}callbackUrl').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}callbackUrl') is not None else ""
                consumerKeyVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}consumerKey').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}consumerKey') is not None else ""
                ipRelaxationVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}ipRelaxation').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}ipRelaxation') is not None else ""
                refreshTokenPolicyVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}refreshTokenPolicy').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}refreshTokenPolicy') is not None else ""
                permissionSetNameVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}permissionSetName').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}permissionSetName') is not None else ""

                scope_items = root.findall('.//{http://soap.sforce.com/2006/04/metadata}scopes')
                scopeText = ", ".join(item.text for item in scope_items if item.text is not None)

                ws.append([labelVar, contactEmailVar, callbackUrlVar, consumerKeyVar, scopeText, ipRelaxationVar, refreshTokenPolicyVar, permissionSetNameVar])

    # Format table
    if ws.max_row > 1:
        tab = Table(displayName=worksheet_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for row in ws.iter_rows():
            for cell in row:
                ws[str(cell.coordinate)].font = Font(size=9)
            for cell in ws["1:1"]:
                cell.font = Font(color='FFFFFF', size=9)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 0.98

        # Save the workbook
        wb.save(utils.config_file_path)
        print(f"Workbook Saved: {utils.config_file_path}")
    else:
        print(f"No current data for this metadata type in {utils.project_path}")
    print(f"Processing complete: {worksheet_name}\n")

# Call the function to execute the script
process_connected_apps()
