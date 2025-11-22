import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import utils
import logging

def process_profiles():
    metadata_folder = 'profiles'
    worksheet_name = metadata_folder
    processing_path = f"{utils.metadata_path}/{metadata_folder}"
    print(f"Processing directory_path {processing_path}")

    if not os.path.exists(processing_path):
        message = f"The directory {processing_path} does not exist. Skipping processing for {worksheet_name}.\n"
        logging.info(message)
        return

    wb = utils.open_workbook(utils.config_file_path)

    if worksheet_name in wb.sheetnames:
        wb.remove(wb[worksheet_name])

    ws = wb.create_sheet(title=worksheet_name)
    header = [
        'name',
        'userLicense',
        'custom',
        'objectPermissions'
    ]
    ws.append(header)
    wb.save(utils.config_file_path)

    for root_dir, dirs, files in os.walk(processing_path):
        i=1
        for file in files:
            i+=1
            file_path = os.path.join(root_dir, file)
            #print(f"Processing file: {file}")
            tree = ET.parse(file_path)
            root = tree.getroot()

            name_var = file.replace('.profile-meta.xml', '')
            userLicenseVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}userLicense').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}userLicense') is not None else ""
            customVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}custom').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}custom') is not None else ""

            ws.cell(row=i,column=1).value = name_var
            ws.cell(row=i,column=2).value = userLicenseVar
            ws.cell(row=i,column=3).value = customVar

            objectPermissions = root.findall('.//{http://soap.sforce.com/2006/04/metadata}objectPermissions')
            j = 0
            objectPermissions_text = ""
            for item in objectPermissions:
                j += 1
                if j > 1:
                    objectPermissions_text += '\n'
                if (object := item.find('.//{http://soap.sforce.com/2006/04/metadata}object')) is not None:
                    objectPermissions_text += f'object: {object.text}, '
                if (allowRead := item.find('.//{http://soap.sforce.com/2006/04/metadata}allowRead')) is not None:
                    objectPermissions_text += f'allowRead: {allowRead.text}, '
                if (allowCreate := item.find('.//{http://soap.sforce.com/2006/04/metadata}allowCreate')) is not None:
                    objectPermissions_text += f'allowCreate: {allowCreate.text}, '
                if (allowEdit := item.find('.//{http://soap.sforce.com/2006/04/metadata}allowEdit')) is not None:
                    objectPermissions_text += f'allowEdit: {allowEdit.text}, '
                if (allowDelete := item.find('.//{http://soap.sforce.com/2006/04/metadata}allowDelete')) is not None:
                    objectPermissions_text += f'allowDelete: {allowDelete.text}, '
                if (viewAllRecords := item.find('.//{http://soap.sforce.com/2006/04/metadata}viewAllRecords')) is not None:
                    objectPermissions_text += f'viewAllRecords: {viewAllRecords.text}, '
                if (modifyAllRecords := item.find('.//{http://soap.sforce.com/2006/04/metadata}modifyAllRecords')) is not None:
                    objectPermissions_text += f'modifyAllRecords: {modifyAllRecords.text}'

            # Step 6
            ws.cell(row=i,column=4).value = objectPermissions_text

        for cell in ws['D']:
            cell.alignment = Alignment(wrap_text=True)

    if ws.max_row > 1:
        tab = Table(displayName=worksheet_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for row in ws.iter_rows():
            for cell in row:
                ws[str(cell.coordinate)].font = Font(size=9)
            for cell in ws["1:1"]:
                cell.font = Font(color='FFFFFF', size=9)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 0.98

        ws.column_dimensions['D'].width = 130

        wb.save(utils.config_file_path)
        print(f"Workbook Saved: {utils.config_file_path}")
    else:
        print(f"No current data for this metadata type in {utils.project_path}")
    print(f"Processing complete: {worksheet_name}\n")

process_profiles()
