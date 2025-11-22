import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import utils
import logging
import re

def process_triggers():
    metadata_folder = 'triggers'
    worksheet_name = metadata_folder
    processing_path = f"{utils.metadata_path}/{metadata_folder}"
    print(f"Processing directory_path {processing_path}")

    if not os.path.exists(processing_path):
        message = f"The directory {processing_path} does not exist. Skipping processing for {worksheet_name}.\n"
        logging.info(message)
        return

    wb = utils.open_workbook(utils.config_file_path)

    if worksheet_name in wb.sheetnames:
        wb.remove(wb[worksheet_name])

    ws = wb.create_sheet(title=worksheet_name)
    header = ['APIName', 'Status', 'ApiVersion', 'Object']
    ws.append(header)
    wb.save(utils.config_file_path)

    # get the .trigger files
    trigger_files = []
    for root, _, files in os.walk(processing_path):
        for file in files:
            if file.endswith(".trigger") and not file.endswith(".trigger-meta"):
                trigger_files.append(os.path.join(root, file))

    # Process .trigger files
    for trigger_file in trigger_files:
        with open(trigger_file, "r") as file:
            content = file.read()
            api_name = os.path.basename(trigger_file).replace('.trigger', '')
            object_name_match = re.search(rf"trigger\s+{api_name}\s+on\s+(\w+)", content)
            object_name = object_name_match.group(1).split("(")[0] if object_name_match else ""

            ws.append([api_name, "", "", object_name])

    # Get the .trigger-meta files
    trigger_meta_files = [os.path.join(root, file) for root, _, files in os.walk(processing_path) for file in files if
                            file.endswith(".trigger-meta.xml")]

    # Process .trigger-meta.xml files
    for trigger_meta_file in trigger_meta_files:
        api_name_var = os.path.basename(trigger_meta_file)
        if api_name_var.endswith(".trigger-meta.xml"):
            api_name_var = api_name_var[:-len(".trigger-meta.xml")]
        elif api_name_var.endswith(".trigger-meta"):
            api_name_var = api_name_var[:-len(".trigger-meta")]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == api_name_var:
                    active_row = cell.row
                    break

        tree = ET.parse(trigger_meta_file)
        root = tree.getroot()
        status_element = root.find(".//{http://soap.sforce.com/2006/04/metadata}status")
        api_version_element = root.find(".//{http://soap.sforce.com/2006/04/metadata}apiVersion")

        if status_element is not None:
            ws.cell(row=active_row, column=2, value=status_element.text)

        if api_version_element is not None:
            ws.cell(row=active_row, column=3, value=api_version_element.text)

    if ws.max_row > 1:
        tab = Table(displayName=worksheet_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for row in ws.iter_rows():
            for cell in row:
                ws[str(cell.coordinate)].font = Font(size=9)
            for cell in ws["1:1"]:
                cell.font = Font(color='FFFFFF', size=9)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 0.98

        wb.save(utils.config_file_path)
        print(f"Workbook Saved: {utils.config_file_path}")
    else:
        print(f"No current data for this metadata type in {utils.project_path}")
    print(f"Processing complete: {worksheet_name}\n")

process_triggers()
