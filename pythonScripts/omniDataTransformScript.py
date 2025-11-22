import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import utils
import logging

def process_omni_data_transform():
    # Metadata folder to process
    metadata_folder = 'omniDataTransforms'
    worksheet_name = metadata_folder

    # Construct the full path to the directory
    processing_path = f"{utils.metadata_path}/{metadata_folder}"
    print(f"Processing directory_path {processing_path}")

    # Check if the directory exists
    if not os.path.exists(processing_path):
        message = f"The directory {processing_path} does not exist. Skipping processing for {worksheet_name}.\n"
        logging.info(message)
        return

    # Get the config file
    wb = utils.open_workbook(utils.config_file_path)

    # Remove sheet if exists
    if worksheet_name in wb.sheetnames:
        wb.remove(wb[worksheet_name])

    # Create new sheet
    ws = wb.create_sheet(title=worksheet_name)

    # Create Header
    header = ['name','active','type','uniqueName','inputAndOutput','description']
    ws.append(header)
    wb.save(utils.config_file_path)

    # Loop through each file in the directory
    for root_dir, dirs, files in os.walk(processing_path):
        i=1
        for file in files:

            i+=1
            tree = ET.parse(os.path.join(root_dir, file))
            root = tree.getroot()

            # Step 5
            nameVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}name').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}name') is not None else ""
            activeVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}active').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}active') is not None else ""
            typeVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}type').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}type') is not None else ""
            uniqueNameVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}uniqueName').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}uniqueName') is not None else ""
            descriptionVar = root.find('.//{http://soap.sforce.com/2006/04/metadata}description').text if root.find('.//{http://soap.sforce.com/2006/04/metadata}description') is not None else ""

            ws.cell(row=i,column=1).value = nameVar
            ws.cell(row=i,column=2).value = activeVar
            ws.cell(row=i,column=3).value = typeVar
            ws.cell(row=i,column=4).value = uniqueNameVar

            omniDataTransformItems = root.findall('.//{http://soap.sforce.com/2006/04/metadata}omniDataTransformItem')
            j = 0
            transform_items_value = ""
            for item in omniDataTransformItems:
                j += 1
                if j > 1:
                    transform_items_value += '\n'
                if (input_object := item.find('.//{http://soap.sforce.com/2006/04/metadata}inputObjectName')) is not None:
                    transform_items_value += f'Input Object: {input_object.text}, '
                if (input_field := item.find('.//{http://soap.sforce.com/2006/04/metadata}inputFieldName')) is not None:
                    transform_items_value += f'Input Field: {input_field.text}, '
                if (output_object := item.find('.//{http://soap.sforce.com/2006/04/metadata}outputObjectName')) is not None:
                    transform_items_value += f'Output Object: {output_object.text}, '
                if (output_field := item.find('.//{http://soap.sforce.com/2006/04/metadata}outputFieldName')) is not None:
                    transform_items_value += f'Output Field: {output_field.text}'

            # Step 6
            ws.cell(row=i,column=5).value = transform_items_value
            ws.cell(row=i,column=6).value = descriptionVar

    for cell in ws['E']:
        cell.alignment = Alignment(wrap_text=True)

    # Format table
    if ws.max_row > 1:
        tab = Table(displayName=worksheet_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for row in ws.iter_rows():
            for cell in row:
                ws[str(cell.coordinate)].font = Font(size=9)
            for cell in ws["1:1"]:
                cell.font = Font(color='FFFFFF', size=9)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 0.98

        # Save the workbook
        wb.save(utils.config_file_path)
        print(f"Workbook Saved: {utils.config_file_path}")
    else:
        print(f"No current data for this metadata type in {utils.project_path}")
    print(f"Processing complete: {worksheet_name}\n")

# Call the function to execute the script
process_omni_data_transform()
