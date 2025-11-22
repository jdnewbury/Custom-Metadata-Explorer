import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import utils
import logging

def process_list_views():
    # Metadata folder to process
    metadata_folder = 'objects'
    worksheet_name = 'listViews'

    # Construct the full path to the directory
    processing_path = f"{utils.metadata_path}/{metadata_folder}"
    print(f"Processing directory_path {processing_path}")

    # Check if the directory exists
    if not os.path.exists(processing_path):
        message = f"The directory {processing_path} does not exist. Skipping processing for {worksheet_name}.\n"
        logging.info(message)
        return

    # Get the config file
    wb = utils.open_workbook(utils.config_file_path)

    # Remove sheet if exists
    if worksheet_name in wb.sheetnames:
        wb.remove(wb[worksheet_name])

    # Create new sheet
    ws = wb.create_sheet(title=worksheet_name)

    # Create Header
    header = ['object', 'fullName', 'filterScope', 'label', 'columns', 'filter1', 'filter2', 'filter3', 'filter4', 'filter5', 'filter6', 'filter7', 'filter8', 'filter9', 'filter10']
    ws.append(header)
    wb.save(utils.config_file_path)

    # Loop through each file in the directory
    for dirpath, dirs, files in os.walk(processing_path):
        for file in files:
            if 'listView-meta' in file:
                file_path = os.path.join(dirpath, file)
                tree = ET.parse(file_path)
                xml_root = tree.getroot()

                # Extract data
                row = [os.path.basename(os.path.dirname(os.path.dirname(file_path)))]  # object
                for tag in ['fullName', 'filterScope', 'label']:
                    element = xml_root.find(f"{{{xml_root.tag.split('}')[0].strip('{')}}}{tag}")
                    row.append(element.text if element is not None else '')
                columns = [e.text for e in xml_root.findall(f"{{{xml_root.tag.split('}')[0].strip('{')}}}columns")]
                row.append(', '.join(columns))

                # Extract filters
                filter_nodes = xml_root.findall(".//{http://soap.sforce.com/2006/04/metadata}filters")
                for filter_node in filter_nodes:
                    field_node = filter_node.find(".//{http://soap.sforce.com/2006/04/metadata}field")
                    operation_node = filter_node.find(".//{http://soap.sforce.com/2006/04/metadata}operation")
                    value_node = filter_node.find(".//{http://soap.sforce.com/2006/04/metadata}value")

                    if value_node is not None:
                        row.append(f"{field_node.text} {operation_node.text} {value_node.text}")

                ws.append(row)

    # Format table
    if ws.max_row > 1:
        tab = Table(displayName=worksheet_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for row in ws.iter_rows():
            for cell in row:
                ws[str(cell.coordinate)].font = Font(size=9)
            for cell in ws["1:1"]:
                cell.font = Font(color='FFFFFF', size=9)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 0.98

        # Save the workbook
        wb.save(utils.config_file_path)
        print(f"Workbook Saved: {utils.config_file_path}")
    else:
        print(f"No current data for this metadata type in {utils.project_path}")
    print(f"Processing complete: {worksheet_name}\n")

# Call the function to execute the script
process_list_views()
