# Configuration module to store shared variables and functions
import os
import logging
import tkinter as tk # for providing GUI elements to the user
from tkinter import filedialog # folder selection
from openpyxl import Workbook, load_workbook #working with Excel
import win32com.client as win32 # type: ignore # for opening the workbook in windows
import datetime

def configure_logging(log_file="../appLogging.log"):
    ### Configures the logging settings for the application. ###
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )

def select_project_directory():
    ### Prompt the user to select a directory and set global configuration variables.###
    global project_path, project_folder, metadata_path, config_matrix_path, config_file_name

    print(f"Please select the project directory using the explorer window that opened...\n")

    root = tk.Tk()
    root.withdraw()  # Hide the root window
    project_path = filedialog.askdirectory(title="Select the VSCode Project Directory\n")

    if not project_path:
        raise ValueError("No project directory selected. Please select a valid directory.\n")

    project_folder = os.path.basename(project_path)
    metadata_path = f"{project_path}/force-app/main/default"
    config_matrix_path = f"{project_path}/config-matrix"
    config_file_name = f'{project_folder} Config Matrix {datetime.datetime.now().strftime("%Y-%m-%d")}.xlsx'

    return

def handle_config_file(config_matrix_path, config_file_name):
    global config_file_path

    # Construct the full path
    config_file_path = os.path.join(config_matrix_path, config_file_name)

    # Ensure the directory exists
    os.makedirs(config_matrix_path, exist_ok=True)

    # Create a new workbook
    wb = Workbook()

    try:
        # Save the workbook to the specified path
        wb.save(config_file_path)
        logging.info(f"Workbook saved successfully at {config_file_path}.\n")
    except PermissionError:
        raise PermissionError(f"Cannot save the workbook because the file {config_file_path} is currently open. Please close it and try again.")

    return

def remove_default_sheet(config_file_path):
    if os.path.exists(config_file_path):
        #print("The config file exists")
        wb = load_workbook(config_file_path)
        if 'Sheet' in wb.sheetnames:
            print(f"removing default sheet\n")
            wb.remove(wb['Sheet'])
            wb.save(config_file_path)
    return

def open_workbook(config_file_path):
    if not os.path.exists(config_file_path):
        raise FileNotFoundError(f"The file {config_file_path} does not exist.\n")

    try:
        logging.info(f"Opening workbook at {config_file_path}")
        return load_workbook(config_file_path)
    except PermissionError:
        raise PermissionError(f"The file {config_file_path} is currently open. Please close it and try again.\n")

def open_excel_visibly(config_file_path):

    ### Opens an Excel file visibly using Excel application. ###

    if not os.path.exists(config_file_path):
        raise FileNotFoundError(f"The file {config_file_path} does not exist.")

    try:
        logging.info(f"Opening Excel visibly for {config_file_path}")
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = True
        workbook = excel.Workbooks.Open(config_file_path)
        return workbook
    except PermissionError:
        raise PermissionError(f"The file {config_file_path} is currently open. Please close it and try again.\n")