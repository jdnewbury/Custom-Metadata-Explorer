import os
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import utils
import logging

def process_assignment_rules():
    metadata_folder = 'assignmentRules'
    worksheet_name = metadata_folder
    processing_path = f"{utils.metadata_path}/{metadata_folder}"
    print(f"Processing directory_path {processing_path}")

    # Check if the directory exists
    if not os.path.exists(processing_path):
        message = f"The directory {processing_path} does not exist. Skipping processing for {worksheet_name}.\n"
        logging.info(message)
        return

    wb = utils.open_workbook(utils.config_file_path)

    if worksheet_name in wb.sheetnames:
        wb.remove(wb[worksheet_name])

    ws = wb.create_sheet(title=worksheet_name)
    header = ['Object', 'FullName', 'Active', 'AssignedTo', 'AssignedToType', 'EmailTemplate', 'Criteria1', 'Criteria2', 'Criteria3', 'Criteria4', 'Criteria5', 'Criteria6', 'Criteria7']
    ws.append(header)
    wb.save(utils.config_file_path)

    for file_name in os.listdir(processing_path):
        file_path = os.path.join(processing_path, file_name)
        if os.path.isfile(file_path):
            #print(f"Processing file: {file_name}")
            process_file(file_path, ws)

    format_worksheet(ws)
    wb.save(utils.config_file_path)
    print(f"Workbook Saved: {utils.config_file_path}")
    print(f"Processing complete: {worksheet_name}\n")

def process_file(file_path, ws):
    object_name = os.path.basename(file_path).split('.')[0]
    tree = ET.parse(file_path)
    root = tree.getroot()
    assignmentRule_items = root.findall('.//{http://soap.sforce.com/2006/04/metadata}assignmentRule')

    for assignmentRule in assignmentRule_items:
        full_name = assignmentRule.find('.//{http://soap.sforce.com/2006/04/metadata}fullName').text if assignmentRule.find('.//{http://soap.sforce.com/2006/04/metadata}fullName') is not None else ''
        active = assignmentRule.find('.//{http://soap.sforce.com/2006/04/metadata}active').text if assignmentRule.find('.//{http://soap.sforce.com/2006/04/metadata}active') is not None else ''

        ruleEntry_items = assignmentRule.findall('.//{http://soap.sforce.com/2006/04/metadata}ruleEntry')
        for ruleEntry in ruleEntry_items:
            assigned_to = ruleEntry.find('.//{http://soap.sforce.com/2006/04/metadata}assignedTo').text if ruleEntry.find('.//{http://soap.sforce.com/2006/04/metadata}assignedTo') is not None else ''
            assigned_to_type = ruleEntry.find('.//{http://soap.sforce.com/2006/04/metadata}assignedToType').text if ruleEntry.find('.//{http://soap.sforce.com/2006/04/metadata}assignedToType') is not None else ''
            email_template = ruleEntry.find('.//{http://soap.sforce.com/2006/04/metadata}template').text if ruleEntry.find('.//{http://soap.sforce.com/2006/04/metadata}template') is not None else ''

            criteria_items = ruleEntry.findall('.//{http://soap.sforce.com/2006/04/metadata}criteriaItems')
            criteria = []
            for item in criteria_items:
                field = item.find('{http://soap.sforce.com/2006/04/metadata}field').text
                operation = item.find('{http://soap.sforce.com/2006/04/metadata}operation').text
                value = item.find('{http://soap.sforce.com/2006/04/metadata}value').text
                criteria_column_value = f"{field} {operation} {value}"
                criteria.append(criteria_column_value)

            row = [object_name, full_name, active, assigned_to, assigned_to_type, email_template] + criteria
            ws.append(row)

def format_worksheet(ws):
    if ws.max_row > 1:
        tab = Table(displayName=ws.title, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for row in ws.iter_rows():
            for cell in row:
                ws[str(cell.coordinate)].font = Font(size=9)
            for cell in ws["1:1"]:
                cell.font = Font(color='FFFFFF', size=10)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
    else:
        print(f"There is currently no data for this metadata type in {utils.project_path}. You'll need to add the type to your manifest file")

# Call the function to execute the script
process_assignment_rules()
